from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

from .content_library import forever_manifesto, workout_library, nutrition_sections

HEADER_FILL = PatternFill("solid", fgColor="D9E1F2")

def _style_header(ws, row=1):
    for cell in ws[row]:
        cell.font = Font(bold=True)
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(wrap_text=True, vertical="top")

def _autosize(ws, widths: dict[int, int]):
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w

def write_master_workbook(out_path: Path, inputs, analysis, day_plans, weekly_rows, narrative: dict | None = None):
    wb = Workbook()

    # Calendar
    ws = wb.active
    ws.title = "Calendar"
    ws.append(["☐","Date","Day","Week#","Phase","Flags","Workout","Notes"])
    _style_header(ws, 1)
    for dp in day_plans:
        ws.append(["", dp.iso_date, dp.day, dp.week, dp.phase, dp.flags, dp.workout, dp.notes])
    ws.freeze_panes = "A2"
    _autosize(ws, {1:4,2:12,3:12,4:7,5:20,6:18,7:20,8:70})

    # Metrics
    ws = wb.create_sheet("Metrics")
    ws.append(list(weekly_rows[0].keys()) if weekly_rows else ["Week#"])
    _style_header(ws, 1)
    for r in weekly_rows:
        ws.append(list(r.values()))
    ws.freeze_panes = "A2"
    _autosize(ws, {1:7,2:20,3:22,4:9,5:13,6:16,7:14,8:16,9:18})

    # Analysis
    ws = wb.create_sheet("Garmin Analysis")
    ws["A1"] = "Garmin Data Analysis Summary"
    ws["A1"].font = Font(size=16, bold=True)
    ws["A3"] = f"Notes: {analysis.notes}"
    ws["A3"].alignment = Alignment(wrap_text=True)
    
    # Column A: Labels
    # Column B: Calculated/Observed Values
    # Column C: Used Values (from inputs)
    
    ws["A5"] = "Observed HRmax"
    ws["B5"] = analysis.hrmax_observed or ""
    ws["C5"] = inputs.athlete.hrmax or ""
    
    ws["A6"] = "Robust HRmax (99.5th pct)"
    ws["B6"] = analysis.hrmax_robust or ""
    ws["C6"] = inputs.athlete.hrmax or "" # Using the same input HRmax as the "used" value
    
    ws["A7"] = "Suggested LTHR (conservative)"
    ws["B7"] = analysis.lthr_suggested or ""
    ws["C7"] = inputs.athlete.lthr or ""
    
    ws["A8"] = "Avg weekly hours"
    ws["B8"] = analysis.avg_weekly_hours or ""
    
    ws["A9"] = "Avg weekly miles"
    ws["B9"] = analysis.avg_weekly_miles or ""
    
    ws["A10"] = "Z2 fraction"
    ws["B10"] = f"{analysis.z2_fraction:.0%}" if analysis.z2_fraction is not None else ""
    
    # Add headers for clarity
    ws["B4"] = "Calculated"
    ws["B4"].font = Font(bold=True)
    ws["C4"] = "Used in Plan"
    ws["C4"].font = Font(bold=True)

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 20

    # Forever Plan
    ws = wb.create_sheet("Forever Plan")
    ws["A1"] = "Forever Training System – Core"
    ws["A1"].font = Font(size=16, bold=True)
    ws["A2"] = f"Athlete: {inputs.athlete.athlete_name} | Age: {inputs.athlete.age} | Event: {inputs.event.event_name} ({inputs.event.event_date})"
    ws["A2"].alignment = Alignment(wrap_text=True)
    ws["A4"] = "Manifesto"
    ws["A4"].font = Font(size=13, bold=True)
    r = 5
    for line in forever_manifesto():
        ws[f"A{r}"] = f"• {line}"
        ws[f"A{r}"].alignment = Alignment(wrap_text=True)
        r += 1
    ws.column_dimensions["A"].width = 100

    # Workout Library
    ws = wb.create_sheet("Workout Library")
    ws["A1"] = "Workout Library (Intervals.icu → Garmin)"
    ws["A1"].font = Font(size=16, bold=True)
    ws["A2"] = "Minimal set of reusable workouts with targets."
    ws["A2"].alignment = Alignment(wrap_text=True)
    r = 4
    for name, bullets in workout_library(inputs.athlete.lthr):
        ws[f"A{r}"] = name
        ws[f"A{r}"].font = Font(size=13, bold=True)
        r += 1
        for b in bullets:
            ws[f"A{r}"] = f"• {b}"
            ws[f"A{r}"].alignment = Alignment(wrap_text=True)
            r += 1
        r += 1
    ws.column_dimensions["A"].width = 100

    # Nutrition
    ws = wb.create_sheet("Nutrition")
    ws["A1"] = "Nutrition & Hydration"
    ws["A1"].font = Font(size=16, bold=True)
    r = 3
    for title, text in nutrition_sections(inputs.athlete.sodium_mg_per_hr_hot):
        ws[f"A{r}"] = title
        ws[f"A{r}"].font = Font(size=13, bold=True)
        r += 1
        ws[f"A{r}"] = text
        ws[f"A{r}"].alignment = Alignment(wrap_text=True)
        r += 2
    ws.column_dimensions["A"].width = 100

    # Narrative (optional)
    if narrative:
        ws = wb.create_sheet("Narrative")
        ws["A1"] = "Narrative Outputs (LLM / Template)"
        ws["A1"].font = Font(size=16, bold=True)
        r = 3
        for k, v in narrative.items():
            ws[f"A{r}"] = k
            ws[f"A{r}"].font = Font(size=13, bold=True)
            r += 1
            ws[f"A{r}"] = str(v)
            ws[f"A{r}"].alignment = Alignment(wrap_text=True, vertical="top")
            r += 2
        ws.column_dimensions["A"].width = 110

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    return out_path
